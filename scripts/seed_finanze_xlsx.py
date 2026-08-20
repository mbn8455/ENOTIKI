#!/usr/bin/env python3
"""Crea il workbook iniziale data/finanze.xlsx a partire dai dati storici del sito."""
import datetime
import pathlib

import openpyxl

import xlsx_to_json as x2j

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "data" / "finanze.xlsx"

POSIZIONI_SEED = [
    ("ETF iShares MSCI World", "ETF", "IE00B4L5Y983", 45.2, 42, 50, 50, "PAC mensile, orizzonte 15+ anni"),
    ("ETF Vanguard FTSE All-World", "ETF", "IE00BK5BQT80", 99.8, 90, 110, 40, ""),
    ("Azioni Enel", "Azioni", "IT0003128367", 6.85, 6, 7.5, 100, "Dividendo interessante"),
    ("Oro fisico (lingotti)", "Oro", "", 2050, 1900, 2200, 50, "Copertura patrimoniale"),
    ("Bitcoin", "Crypto", "", 58000, 45000, 70000, 70, ""),
]

MOVIMENTI_SEED_BOOTSTRAP = [
    ("ETF iShares MSCI World", 120, 38.5),
    ("ETF Vanguard FTSE All-World", 60, 95),
    ("Azioni Enel", 200, 6.1),
    ("Oro fisico (lingotti)", 3, 1850),
    ("Bitcoin", 0.15, 42000),
]

DEPOSITI_SEED = [
    ("Marco", 8000, datetime.date(2026, 2, 10), "Bonus annuale"),
    ("Eugenia", 6000, datetime.date(2026, 2, 15), ""),
    ("Marco", 4000, datetime.date(2026, 5, 3), ""),
    ("Eugenia", 5390, datetime.date(2026, 6, 20), "Tredicesima"),
]

BOOTSTRAP_DATE = datetime.date(2026, 8, 12)


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_pos = wb.create_sheet("Posizioni")
    ws_pos.append(x2j.POSIZIONI_HEADERS)
    for nome, tipo, isin, valore_attuale, range_min, range_max, quota_marco, note in POSIZIONI_SEED:
        ws_pos.append([nome, tipo, isin, 0, 0, valore_attuale, range_min, range_max, quota_marco, note])

    ws_mov = wb.create_sheet("Movimenti")
    ws_mov.append(x2j.MOVIMENTI_HEADERS)
    for nome, quantita, prezzo in MOVIMENTI_SEED_BOOTSTRAP:
        ws_mov.append([BOOTSTRAP_DATE, nome, "Acquisto", quantita, prezzo, "Saldo iniziale importato dal sito"])

    ws_dep = wb.create_sheet("Depositi")
    ws_dep.append(x2j.DEPOSITI_HEADERS)
    for persona, importo, data, note in DEPOSITI_SEED:
        ws_dep.append([persona, importo, data, note])

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"Creato {XLSX_PATH}")


if __name__ == "__main__":
    main()
