#!/usr/bin/env python3
"""Rigenera investimenti.json e depositi.json a partire da data/finanze.xlsx."""
import datetime
import json
import pathlib
import re
import sys

import openpyxl

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "data" / "finanze.xlsx"
INVESTIMENTI_JSON = REPO_ROOT / "investimenti.json"
DEPOSITI_JSON = REPO_ROOT / "depositi.json"

POSIZIONI_HEADERS = ["nome", "tipo", "isin", "quantita", "prezzoCarico", "valoreAttuale", "rangeMin", "rangeMax", "quotaMarco", "note"]
MOVIMENTI_HEADERS = ["data", "nome", "tipo", "quantita", "prezzo", "note"]
DEPOSITI_HEADERS = ["persona", "importo", "data", "note"]


def apply_movimento(quantita, costo_totale, movimento):
    if movimento["tipo"] == "Acquisto":
        quantita += movimento["quantita"]
        costo_totale += movimento["quantita"] * movimento["prezzo"]
    elif movimento["tipo"] == "Vendita":
        if quantita > 0:
            prezzo_medio = costo_totale / quantita
            costo_totale -= prezzo_medio * movimento["quantita"]
        quantita -= movimento["quantita"]
    return quantita, costo_totale


def compute_posizione(base, movimenti):
    quantita, costo_totale = 0.0, 0.0
    rilevanti = sorted((m for m in movimenti if m["nome"] == base["nome"]), key=lambda m: m.get("data", ""))
    for m in rilevanti:
        quantita, costo_totale = apply_movimento(quantita, costo_totale, m)
    prezzo_carico = costo_totale / quantita if quantita > 0 else 0.0
    return {**base, "quantita": round(quantita, 6), "prezzoCarico": round(prezzo_carico, 6)}


def _iso_date(value):
    if isinstance(value, (datetime.date, datetime.datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value) if value is not None else ""


def _sheet_rows(ws, headers):
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if header_row[: len(headers)] != headers:
        raise ValueError(f"Foglio '{ws.title}': intestazioni inattese {header_row[: len(headers)]!r}, attese {headers!r}")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def _coerce_numeric(rows, fields, sheet_name):
    for row in rows:
        for field in fields:
            value = row[field]
            try:
                row[field] = float(value)
            except (TypeError, ValueError):
                raise ValueError(f"Foglio '{sheet_name}': valore non numerico per '{field}': {value!r}")
    return rows


def load_workbook_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    posizioni = _sheet_rows(wb["Posizioni"], POSIZIONI_HEADERS)
    movimenti = _sheet_rows(wb["Movimenti"], MOVIMENTI_HEADERS)
    depositi = _sheet_rows(wb["Depositi"], DEPOSITI_HEADERS)
    _coerce_numeric(posizioni, ["valoreAttuale", "rangeMin", "rangeMax", "quotaMarco"], "Posizioni")
    _coerce_numeric(movimenti, ["quantita", "prezzo"], "Movimenti")
    _coerce_numeric(depositi, ["importo"], "Depositi")
    for m in movimenti:
        m["data"] = _iso_date(m["data"])
    for d in depositi:
        d["data"] = _iso_date(d["data"])
    return wb, posizioni, movimenti, depositi


def slugify(nome):
    return re.sub(r"[^a-z0-9]+", "-", nome.lower()).strip("-")


def _write_computed_posizioni(wb, posizioni_computed):
    ws = wb["Posizioni"]
    by_nome = {p["nome"]: p for p in posizioni_computed}
    for row_idx in range(2, ws.max_row + 1):
        nome = ws.cell(row=row_idx, column=1).value
        if nome is None:
            continue
        p = by_nome.get(nome)
        if not p:
            continue
        ws.cell(row=row_idx, column=4).value = p["quantita"]
        ws.cell(row=row_idx, column=5).value = p["prezzoCarico"]


def build_investimenti_json(posizioni_computed):
    return {
        "generatedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "investimenti": [
            {
                "id": "fin-" + slugify(p["nome"]),
                "nome": p["nome"],
                "tipo": p["tipo"],
                "isin": p["isin"] or "",
                "quantita": p["quantita"],
                "prezzoCarico": p["prezzoCarico"],
                "valoreAttuale": p["valoreAttuale"],
                "rangeMin": p["rangeMin"],
                "rangeMax": p["rangeMax"],
                "quotaMarco": p["quotaMarco"],
                "note": p["note"] or "",
            }
            for p in posizioni_computed
        ],
    }


def build_depositi_json(depositi):
    return {
        "generatedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "depositi": [
            {
                "id": "dep-" + str(i),
                "persona": d["persona"],
                "importo": d["importo"],
                "data": d["data"],
                "note": d["note"] or "",
            }
            for i, d in enumerate(depositi, start=1)
        ],
    }


def main():
    wb, posizioni, movimenti, depositi = load_workbook_data(XLSX_PATH)
    posizioni_computed = [compute_posizione(p, movimenti) for p in posizioni]
    _write_computed_posizioni(wb, posizioni_computed)
    wb.save(XLSX_PATH)
    INVESTIMENTI_JSON.write_text(
        json.dumps(build_investimenti_json(posizioni_computed), indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    DEPOSITI_JSON.write_text(
        json.dumps(build_depositi_json(depositi), indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(f"Scritti {INVESTIMENTI_JSON.name} ({len(posizioni_computed)} posizioni) e {DEPOSITI_JSON.name} ({len(depositi)} depositi).")


if __name__ == "__main__":
    sys.exit(main())
