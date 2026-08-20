import pytest

import xlsx_to_json as x2j


def test_apply_movimento_singolo_acquisto():
    quantita, costo = x2j.apply_movimento(0, 0, {"tipo": "Acquisto", "quantita": 10, "prezzo": 5})
    assert quantita == 10
    assert costo == 50


def test_apply_movimento_due_acquisti_media_ponderata():
    quantita, costo = x2j.apply_movimento(0, 0, {"tipo": "Acquisto", "quantita": 100, "prezzo": 10})
    quantita, costo = x2j.apply_movimento(quantita, costo, {"tipo": "Acquisto", "quantita": 50, "prezzo": 16})
    assert quantita == 150
    assert costo == pytest.approx(100 * 10 + 50 * 16)


def test_apply_movimento_vendita_riduce_quantita_mantiene_prezzo_medio():
    quantita, costo = x2j.apply_movimento(0, 0, {"tipo": "Acquisto", "quantita": 100, "prezzo": 10})
    quantita, costo = x2j.apply_movimento(quantita, costo, {"tipo": "Vendita", "quantita": 40, "prezzo": 20})
    assert quantita == 60
    assert costo == pytest.approx(600)  # prezzo medio resta 10, tolti 40*10 di costo


def test_apply_movimento_vendita_totale_azzera_costo():
    quantita, costo = x2j.apply_movimento(0, 0, {"tipo": "Acquisto", "quantita": 10, "prezzo": 5})
    quantita, costo = x2j.apply_movimento(quantita, costo, {"tipo": "Vendita", "quantita": 10, "prezzo": 8})
    assert quantita == 0
    assert costo == pytest.approx(0)


def test_compute_posizione_media_ponderata_su_due_acquisti():
    base = {"nome": "ETF Test", "tipo": "ETF", "isin": "", "valoreAttuale": 10, "rangeMin": 0, "rangeMax": 0, "quotaMarco": 50, "note": ""}
    movimenti = [
        {"nome": "ETF Test", "tipo": "Acquisto", "quantita": 100, "prezzo": 10},
        {"nome": "ETF Test", "tipo": "Acquisto", "quantita": 50, "prezzo": 16},
        {"nome": "Altro Strumento", "tipo": "Acquisto", "quantita": 999, "prezzo": 1},
    ]
    result = x2j.compute_posizione(base, movimenti)
    assert result["nome"] == "ETF Test"
    assert result["quantita"] == 150
    assert result["prezzoCarico"] == pytest.approx((100 * 10 + 50 * 16) / 150)


def test_compute_posizione_quantita_zero_prezzo_carico_zero():
    base = {"nome": "ETF Vuoto", "tipo": "ETF", "isin": "", "valoreAttuale": 10, "rangeMin": 0, "rangeMax": 0, "quotaMarco": 50, "note": ""}
    result = x2j.compute_posizione(base, [])
    assert result["quantita"] == 0
    assert result["prezzoCarico"] == 0


def test_compute_posizione_ordina_per_data_indipendentemente_dall_ordine_lista():
    base = {"nome": "ETF Test", "tipo": "ETF", "isin": "", "valoreAttuale": 10, "rangeMin": 0, "rangeMax": 0, "quotaMarco": 50, "note": ""}
    movimenti_data_order = [
        {"nome": "ETF Test", "data": "2026-01-01", "tipo": "Acquisto", "quantita": 100, "prezzo": 10},
        {"nome": "ETF Test", "data": "2026-02-01", "tipo": "Acquisto", "quantita": 50, "prezzo": 16},
        {"nome": "ETF Test", "data": "2026-03-01", "tipo": "Vendita", "quantita": 40, "prezzo": 12},
    ]
    movimenti_reversed = list(reversed(movimenti_data_order))

    result_ordine_date = x2j.compute_posizione(base, movimenti_data_order)
    result_ordine_inverso = x2j.compute_posizione(base, movimenti_reversed)

    # L'ordine in cui i movimenti arrivano nella lista non deve influire: compute_posizione
    # li riordina internamente per data.
    assert result_ordine_inverso["quantita"] == result_ordine_date["quantita"]
    assert result_ordine_inverso["prezzoCarico"] == pytest.approx(result_ordine_date["prezzoCarico"])
    # Valori attesi calcolati replay-ando in ordine cronologico:
    # Acquisto 100@10 -> qty=100 costo=1000; Acquisto 50@16 -> qty=150 costo=1800;
    # Vendita 40@12 (prezzo medio 12) -> qty=110 costo=1320 -> prezzoCarico=12.0
    assert result_ordine_date["quantita"] == 110
    assert result_ordine_date["prezzoCarico"] == pytest.approx(12.0)


import datetime
import json

import openpyxl


def _build_test_workbook(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_pos = wb.create_sheet("Posizioni")
    ws_pos.append(x2j.POSIZIONI_HEADERS)
    ws_pos.append(["ETF Test", "ETF", "IE000000", 0, 0, 12.0, 8.0, 15.0, 50, "Nota"])
    ws_mov = wb.create_sheet("Movimenti")
    ws_mov.append(x2j.MOVIMENTI_HEADERS)
    ws_mov.append([datetime.date(2026, 1, 1), "ETF Test", "Acquisto", 10, 5, ""])
    ws_mov.append([datetime.date(2026, 2, 1), "ETF Test", "Acquisto", 10, 7, ""])
    ws_dep = wb.create_sheet("Depositi")
    ws_dep.append(x2j.DEPOSITI_HEADERS)
    ws_dep.append(["Marco", 1000, datetime.date(2026, 1, 5), "Nota deposito"])
    wb.save(path)


def test_load_workbook_data_converte_date_in_stringa(tmp_path):
    xlsx_path = tmp_path / "finanze.xlsx"
    _build_test_workbook(xlsx_path)
    wb, posizioni, movimenti, depositi = x2j.load_workbook_data(xlsx_path)
    assert posizioni[0]["nome"] == "ETF Test"
    assert movimenti[0]["data"] == "2026-01-01"
    assert depositi[0]["data"] == "2026-01-05"


def test_load_workbook_data_intestazioni_inattese_solleva_errore(tmp_path):
    xlsx_path = tmp_path / "finanze.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_pos = wb.create_sheet("Posizioni")
    intestazioni_sbagliate = list(x2j.POSIZIONI_HEADERS)
    intestazioni_sbagliate[0] = "nomeStrumento"  # colonna rinominata rispetto all'atteso
    ws_pos.append(intestazioni_sbagliate)
    ws_pos.append(["ETF Test", "ETF", "IE000000", 0, 0, 12.0, 8.0, 15.0, 50, "Nota"])
    ws_mov = wb.create_sheet("Movimenti")
    ws_mov.append(x2j.MOVIMENTI_HEADERS)
    ws_dep = wb.create_sheet("Depositi")
    ws_dep.append(x2j.DEPOSITI_HEADERS)
    wb.save(xlsx_path)

    with pytest.raises(ValueError):
        x2j.load_workbook_data(xlsx_path)


def test_load_workbook_data_valore_non_numerico_solleva_errore(tmp_path):
    xlsx_path = tmp_path / "finanze.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_pos = wb.create_sheet("Posizioni")
    ws_pos.append(x2j.POSIZIONI_HEADERS)
    ws_pos.append(["ETF Test", "ETF", "IE000000", 0, 0, 12.0, 8.0, 15.0, 50, "Nota"])
    ws_mov = wb.create_sheet("Movimenti")
    ws_mov.append(x2j.MOVIMENTI_HEADERS)
    ws_mov.append([datetime.date(2026, 1, 1), "ETF Test", "Acquisto", "dieci", 5, ""])
    ws_dep = wb.create_sheet("Depositi")
    ws_dep.append(x2j.DEPOSITI_HEADERS)
    wb.save(xlsx_path)

    with pytest.raises(ValueError):
        x2j.load_workbook_data(xlsx_path)


def test_main_genera_json_e_riscrive_workbook(tmp_path, monkeypatch):
    xlsx_path = tmp_path / "finanze.xlsx"
    _build_test_workbook(xlsx_path)
    investimenti_json = tmp_path / "investimenti.json"
    depositi_json = tmp_path / "depositi.json"
    monkeypatch.setattr(x2j, "XLSX_PATH", xlsx_path)
    monkeypatch.setattr(x2j, "INVESTIMENTI_JSON", investimenti_json)
    monkeypatch.setattr(x2j, "DEPOSITI_JSON", depositi_json)

    x2j.main()

    inv_data = json.loads(investimenti_json.read_text(encoding="utf-8"))
    assert inv_data["investimenti"][0]["nome"] == "ETF Test"
    assert inv_data["investimenti"][0]["quantita"] == 20
    assert inv_data["investimenti"][0]["prezzoCarico"] == pytest.approx(6.0)
    assert inv_data["investimenti"][0]["id"] == "fin-etf-test"

    dep_data = json.loads(depositi_json.read_text(encoding="utf-8"))
    assert dep_data["depositi"][0]["persona"] == "Marco"
    assert dep_data["depositi"][0]["importo"] == 1000
    assert dep_data["depositi"][0]["id"] == "dep-1"

    wb_after = openpyxl.load_workbook(xlsx_path)
    ws_pos = wb_after["Posizioni"]
    assert ws_pos.cell(row=2, column=4).value == 20
    assert ws_pos.cell(row=2, column=5).value == pytest.approx(6.0)
